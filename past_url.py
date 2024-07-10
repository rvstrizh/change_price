from openpyxl import load_workbook
import pymysql
from config import host, user, password, db_name

connection = pymysql.connect(
    host=host,
    port=3306,
    user=user,
    password=password,
    database=db_name,
    cursorclass=pymysql.cursors.DictCursor
)

class Open_Save:
    def __init__(self):
        self.counter = 0
        global wb
        wb = load_workbook(f'past_url.xlsx')

    def open(self, sheet_name):
        sheet = wb[f'{sheet_name}']
        row = self.counter
        self.products = {}
        while True:
            row += 1
            if sheet.cell(row=row, column=1).value is None:
                break
            self.products[sheet.cell(row=row, column=1).value] = sheet.cell(row=row, column=2).value
        return self.products

    def write_sql(self, url_SMM, product_id):
        with connection.cursor() as cursor:
            sql_query = """
                UPDATE nsrf_vm_product_type_103
                SET url_SMM = %s
                WHERE product_id = %s
            """
            cursor.execute(sql_query, (url_SMM, product_id))
            connection.commit()
            connection.close()

    def run(self):
        products = self.open('product')
        for product, url in products.items():
            self.write_sql(url, product)

if __name__ == "__main__":
    s = Open_Save().run()
    print(s)
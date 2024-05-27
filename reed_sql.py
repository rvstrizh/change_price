from openpyxl import load_workbook


class Open_Save:
    def __init__(self, user):
        self.counter = 0
        self.user = user
        global wb
        wb = load_workbook(f'{self.user}/check_price.xlsx')

    def open(self, sheet_name):
        sheet = wb[f'{sheet_name}']
        row = self.counter
        self.products = []
        while True:
            row += 1
            if sheet.cell(row=row, column=1).value is None:
                break
            self.products.append([sheet.cell(row=row, column=1).value,
                                 sheet.cell(row=row, column=2).value,
                                 sheet.cell(row=row, column=3).value,
                                 sheet.cell(row=row, column=4).value,
                                 sheet.cell(row=row, column=5).value,
                                sheet.cell(row=row, column=6).value])
        return self.products


if __name__ == "__main__":
    s = Open_Save('1315757744').open('Кофемашины')
    print(s)
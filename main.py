import json
import time
import schedule
from datetime import datetime
import xlsxwriter
from aiogram.utils import executor
import threading

from threading import Thread
from bot_change_price import dp
from connect_sql import read_sql, write_sql
from driver import installation
from parse import Parse_Page
from calculation_purchase import Search_Prices_For_Purchase
from calculation_price import Price_Change
from settings import CaptchaError, category_dict, bot, open_json, got_notice, past_site
from openpyxl import load_workbook


def create_file():
    names = ["Roman", "Dmitriy", "Andrey_P", "Sergey", "Andrey_K"]
    with open(f'./not_notice.json', 'w', encoding='utf-8') as f:
        json.dump({}, f)
    for name in names:
        with open(f'./offers_price/{name}/{name}.json', 'w', encoding='utf-8') as f:
            json.dump({}, f)
        name_file = f'./offers_price/{name}/{name}.xlsx'
        workbook = xlsxwriter.Workbook(name_file)
        workbook.add_worksheet('Списание')
        workbook.add_worksheet('Накопление')
        workbook.close()
        wb = load_workbook(name_file)
        for sheet in ['Списание', 'Накопление']:
            sheets = wb[sheet]
            sheets[f'A1'], sheets[f'B1'], sheets[f'C1'], sheets[f'D1'] = \
                "Название", "Ссылка", "Цена", "Разница в процентах"
            wb.save(name_file)


def write_json(price):
    with open('bd_sql.json', 'w') as f:
        json.dump(price, f, indent=4)


schedule.every().day.at("23:50").do(create_file)


class Main:
    def __init__(self):
        self.stock = None
        self.product_name = None
        self.product_id = None
        self.category = None
        self.url_smm = None
        self.old_my_price = None
        self.min_price = None
        self.max_price = None
        self.sensitivity = None
        self.step = None
        self.manager = None
        self.manager_tel_id = None
        # self.user_agent = installation()
        self.user_agent ='Mozilla/5.0 (Linux; Android 10; K) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Mobile Safari/537.36'

    def parse(self, url_smm):
        while True:
            try:
                price_list = Parse_Page(url_smm, self.user_agent).show_shop()
                return price_list
            except CaptchaError:
                self.user_agent = installation()

    def new_price(self, price_list):
        ######
        print(price_list)
        #####
        if 'MegaPixel' not in price_list:
            # поставить сигнал боту что силениум не нашел наш магазин, хотя товар в наличии
            if got_notice(self.product_id, self.product_name):
                bot.send_message(self.manager_tel_id, f'Не нашел наш магазин,в карточке товара sku {self.product_id}, хотя наличие в админке проставлено, убедись точно, вот ссылка:\n{self.url_smm}')

        # Search_Prices_For_Purchase(price_list, self.manager, self.manager_tel_id, self.product_name, self.url_smm,
        #                            self.old_my_price).run()
        # прописать с товаром которого нет в наличии, но нам нужно посмотреть карточку на выгодный кешбек и просто на низкую цену
        return Price_Change(price_list, self.min_price, self.max_price, self.sensitivity, self.step).run()

    def run(self):
        bot.set_webhook()
        # for i in range(1):
        while True:
            current_hour = datetime.now().hour
            if 7 <= current_hour < 24:
                json_file = open_json('bd_sql.json')
                if json_file:
                    self.product_name = json_file[0]['product_sku']
                    self.product_id = json_file[0]['product_id']
                    self.category = json_file[0]['category']
                    self.url_smm = json_file[0]['url_SMM']
                    self.old_my_price = json_file[0]['product_order_levels']
                    self.min_price = json_file[0]['min_price_smm']
                    self.max_price = json_file[0]['max_price_SMM']
                    self.sensitivity = json_file[0]['sensitivity']
                    self.step = json_file[0]['step']
                    self.stock = json_file[0]['product_in_stock']
                    self.old_my_price = int(self.old_my_price.replace("0,", ""))
                    try:
                        self.manager = [man for man, cat in category_dict.items() if self.category in cat][0]
                        self.manager_tel_id = [k for k, v in open_json('data.json').items() if v == self.manager][0]
                    except IndexError:
                        self.manager = 'Roman'
                        self.manager_tel_id = 1315757744 # когда manager_tel_id не задан
                        if got_notice(self.product_id, self.product_name): # что бы не спамить, а уведомление будет один раз в день
                            bot.send_message(1315757744, f'Категория № {self.category} товар\n'
                                                         f'{self.product_name}\n'
                                                         f'не внесен в список категорий менеджеров')
                    if self.url_smm:
                        price_list = self.parse(self.url_smm)
                        if self.stock:
                            if all([self.min_price, self.max_price, self.step]):
                                if self.min_price < self.max_price:
                                    new_price = self.new_price(price_list)
                                    # временное дополнение что бы не менять автоматом цену
                                    if self.old_my_price != new_price:
                                        # if self.category in category_dict['Roman']:
                                        #     write_sql(new_price, self.product_id) # изменение цены в sql
                                        bot.send_message(self.manager_tel_id,
                                                         f'Поменял цену sku {past_site(self.product_id)}\n'
                                                         f'{self.product_name}\n'
                                                         f'на {new_price}, вот тебе ссылка на мегамаркет\n'
                                                         f'{self.url_smm}', parse_mode='HTML')
                                else:
                                    if got_notice(self.product_id,
                                                  self.product_name):  # что бы не спамить, а уведомление будет один раз в день
                                        bot.send_message(self.manager_tel_id, f'В sku {past_site(self.product_id)} {self.product_name}'
                                                                              f'\nминимальная цена, больше максимальной', parse_mode='HTML')
                            else:
                                if got_notice(self.product_id,
                                              self.product_name):  # что бы не спамить, а уведомление будет один раз в день
                                    bot.send_message(self.manager_tel_id, f'В товаре {self.product_name} sku {past_site(self.product_id)}\nmin_price = {self.min_price}\nmax_price = {self.max_price}', parse_mode='HTML')

                        Search_Prices_For_Purchase(price_list, self.manager, self.manager_tel_id, self.product_name,
                                                   self.url_smm,
                                                   self.old_my_price).run()
                    else:
                        bot.set_webhook()
                        if got_notice(self.product_id,
                                      self.product_name):  # что бы не спамить, а уведомление будет один раз в день
                            bot.send_message(self.manager_tel_id, f'В товаре {self.product_name} sku {past_site(self.product_id)}\nНе заполнено url_smm', parse_mode='HTML')
                    write_json(json_file[1:])
                else:
                    bot.send_message(1315757744,
                                     f'Прошел круг')
                    read_sql()

            schedule.run_pending()
            time.sleep(1)


if __name__ == "__main__":
    # print(open_json('data.json'))

    Main().run()
    executor.start_polling(dp, skip_updates=True)
    # create_file()
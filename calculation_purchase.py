import json

from settings import category_dict, bot, open_json
from openpyxl import load_workbook


class Search_Prices_For_Purchase:
    def __init__(self, price_list, manager, manager_tel_id, product_name, url_smm, old_my_price=0):
        self.price_list = list(map(lambda x: x, price_list.values()))
        self.product_name = product_name
        self.manager = manager
        self.manager_tel_id = manager_tel_id
        self.url_smm = url_smm
        self.old_my_price = old_my_price
        self.flag = False
        self.name_file = f'./offers_price/{self.manager}/{self.manager}.xlsx'
        self.wb = load_workbook(self.name_file)

    def save_price(self, min_price, sheet, difference_price):
        with open(f'./offers_price/{self.manager}/{self.manager}.json', 'r') as f:
            data = json.load(f)
        try:
            data[self.product_name]
        except KeyError:
            data[self.product_name] = min_price
            with open(f'./offers_price/{self.manager}/{self.manager}.json', 'w') as f:
                json.dump(data, f)
            sheets = self.wb[sheet]
            last_filled_row = sheets.max_row + 1
            sheets[f'A{last_filled_row}'], sheets[f'B{last_filled_row}'], sheets[f'C{last_filled_row}'], sheets[f'D{last_filled_row}'] = \
                self.product_name, self.url_smm, min_price, difference_price
            self.wb.save(self.name_file)

    def search_for_accrual(self):  # буду возвращать слово 'Накопление'
        min_price = min(price[0] for price in self.price_list)
        best_price = min(price[0] * 0.96 - (price[1] * 0.7) for price in self.price_list)
        # если минимальная минус 25% больше чем цена минимальная тогда эта карточка мне подходит
        if min_price * 0.75 > best_price:
            difference_price = best_price * 100 / min_price
            self.save_price(min_price, 'Накопление', difference_price)
            # bot.send_message(self.manager_tel_id, text=f'Накопление\n<a href="{self.url_smm}">{self.product_name}</a>', parse_mode='HTML')

# почему то на списание подбирает включая мое предложение и игнорирует самовывоз и доставку продавцом магазина
    def search_for_write_off(self):  # буду возвращать слово 'Списание'
        min_price = min(price[0] for price in self.price_list)
        sorted_price_list = sorted(self.price_list, key=lambda x: x[0])
        if self.old_my_price > 2:
            if self.old_my_price * 0.8 > min_price:
                difference_price = min_price * 100 / self.old_my_price
                self.save_price(min_price, 'Списание', difference_price)
                # bot.send_message(self.manager_tel_id, text=f'Списание\n<a href="{self.url_smm}">{self.product_name}</a>', parse_mode='HTML')
            elif sorted_price_list[1][0] * 0.8 > sorted_price_list[0][0]:
                difference_price = sorted_price_list[0][0] * 100 / sorted_price_list[1][0]
                self.save_price(min_price, 'Списание', difference_price)
                # bot.send_message(self.manager_tel_id, text=f'Списание\n<a href="{self.url_smm}">{self.product_name}</a>', parse_mode='HTML')
        else:
            if min_price < sorted_price_list[1][0] * 0.8:
                difference_price = min_price * 100 / sorted_price_list[1][0]
                self.save_price(min_price, 'Списание', difference_price)

    def run(self):
        if len(self.price_list) > 1:
            self.search_for_accrual()
            self.search_for_write_off()
            return self.flag
        else:
            return False


if __name__ == "__main__":
    sp = Search_Prices_For_Purchase({'MegaPixel': [57990, 1160, 'Курьером СММ'], 'Just a store': [87980, 1760, 'Курьером СММ']}, 'Roman', 1315757744, 'Huawei MatePad Pro 11.0 LTE 8/256Gb Gray (RU)', 'https://megamarket.ru/catalog/details/planshet-huawei-matepad-pro-11-8-256gb-lte-goethe-al09bs-golden-black-53013gak-100044896674/#?details_block=prices').run()
    print(sp)